# companion_app.py
# Author:Josh Smith

import sys
import tkinter as tk
from tkinter import ttk
from tkcalendar import Calendar
from tkcalendar import DateEntry
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class BoundText(tk.Text):
    """A Text widget with a bound variable
    Credit: Alan D. Moore "Python GUI Programming with Tkinter"""

    def __init__(self, *args, textvariable=None, **kwargs):
        super().__init__(*args, **kwargs)
        self._variable = textvariable
        if self._variable:
            self.insert('1.0', self._variable.get())
            self.bind('<<Modified>>', self._set_var)

    def _set_content(self, *_):
        """Set the text contents to the variable"""
        self.delete('1.0', tk.END)
        self.insert('1.0', self._variable.get())

    def _set_var(self, *_):
        """Set the variable to the text contents"""
        if self.edit_modified():
            content = self.get('1.0', 'end-1chars')
            self._variable.set(content)
            self.edit_modified(False)


class LabelInput(tk.Frame):
    """A widget containing a label and input together.
    Credit: Alan D. Moore "Python GUI Programming with Tkinter"""

    def __init__(
            self, parent, label, var, input_class=ttk.Entry,
            input_args=None, label_args=None, **kwargs
    ):
        super().__init__(parent, **kwargs)
        input_args = input_args or {}
        label_args = label_args or {}
        # The above statements say if label_args or input_args are not None,
        # they are what was passed during init.
        # However, if they are None, then make them empty dicts
        self.variable = var
        self.variable.label_widget = self

        if input_class in (ttk.Checkbutton, ttk.Button):
            input_args["text"] = label
        else:
            self.label = ttk.Label(self, text=label, **label_args)
            self.label.grid(row=0, column=0, sticky=(tk.W + tk.E))

        if input_class in (
            ttk.Checkbutton, ttk.Button, ttk.Radiobutton
        ):
            input_args["variable"] = self.variable
        else:
            input_args["textvariable"] = self.variable

        if input_class == ttk.Radiobutton:
            self.input = tk.Frame(self)
            for v in input_args.pop('values', []):
                button = ttk.Radiobutton(
                    self.input, value=v, text=v, **input_args
                )
                button.pack(
                    side=tk.LEFT, ipadx=10, ipady=2, expand=True, fill='x'
                )
        else:
            self.input = input_class(self, **input_args)

        self.input.grid(row=1, column=0, sticky=(tk.E + tk.W))
        self.columnconfigure(0, weight=1)

    def grid(self, sticky=(tk.E + tk.W), **kwargs):
        """Override grid to add default sticky values"""
        super().grid(sticky=sticky, **kwargs)


class AppPage(ttk.Frame):
    """Application page class from which all other pages will inherit."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._vars = {
            'Client Name': tk.StringVar(),
            'Go Live / Dead Date': tk.StringVar(),
            'Market': tk.StringVar(),
            'Type': tk.StringVar(),
            'AV': tk.StringVar(),
            'MDR': tk.StringVar(),
            'Ninjio': tk.StringVar(),
            'Barracuda': tk.StringVar(),
            'AV-conv': tk.StringVar(),
            'MDR-conv': tk.StringVar(),
            'Ninjio-conv': tk.StringVar(),
            'Barracuda-conv': tk.StringVar(),
            'update1_choice': tk.StringVar(),
        }

    def _add_frame(self, label, cols=2):
        """Add a label frame to the form
        Credit: Alan D. Moore "Python GUI Programming with Tkinter"""

        frame = ttk.LabelFrame(self, text=label)
        frame.grid(sticky=tk.W + tk.E)
        for i in range(cols):
            frame.columnconfigure(i, weight=1)
        return frame

    def get(self):
        data = dict()
        for key, variable in self._vars.items():
            try:
                data[key] = variable.get()
            except tk.TclError:
                message = f'Error in field: {key}. Data was not saved!'
                raise ValueError(message)
        return data

    def _tpg_tools(self, p_type, av, mdr, barracuda, ninjio):
        """Lays out the frames needed to select TPG tools.
        This is needed for onbaordings(to see what to deploy)
        Offboardings(to see what to pull)
        Conversiones(to see what to pull before push or just review)"""

        if p_type == 'Offboarding':
            av_info = self._add_frame(f"Which AV are we pulling?")
            mdr_info = self._add_frame(f"Which MDR are we pulling?")
            ess_info = self._add_frame(
                f"Which Barracuda level are we pulling?")
            ninjio_info = self._add_frame(
                "Do they currently get Security Training?"
            )
        elif p_type == 'Conversion':
            av_info = self._add_frame(f"Which AV do they currently have?")
            mdr_info = self._add_frame(f"Which MDR do they currently have?")
            ess_info = self._add_frame(
                f"Which Barracuda level do they currently have?")
            ninjio_info = self._add_frame(
                "Do they currently get Security Training?"
            )
        else:
            av_info = self._add_frame(f"Which AV are they getting?")
            mdr_info = self._add_frame(f"Which MDR are they getting?")
            ess_info = self._add_frame(
                f"Which Barracuda level are they getting?")
            ninjio_info = self._add_frame(
                "Are they getting Security Training?"
            )

        LabelInput(
            av_info, "", input_class=ttk.Radiobutton,
            var=self._vars[av],
            input_args={"values": ['Sophos', 'AV Defender', 'None']}
        ).grid(row=0, column=0)

        LabelInput(
            mdr_info, "", input_class=ttk.Radiobutton,
            var=self._vars[mdr],
            input_args={"values": ['Blackpoint', 'Arctic Wolf', 'None']}
        ).grid(row=0, column=0)

        LabelInput(
            ess_info, "", input_class=ttk.Radiobutton,
            var=self._vars[barracuda],
            input_args={"values": ['Complete+Sentinel',
                                   'Complete Only',
                                   'Spam only',
                                   'None'
                                   ]
                        }
        ).grid(row=0, column=0)

        LabelInput(
            ninjio_info, "Ninjio", input_class=ttk.Checkbutton,
            var=self._vars[ninjio],
        ).grid(row=0, column=0)

    def create_title(self, client, p_type):
        client_label = ttk.Label(
            self,
            text=f"{client}'s {p_type}",
            font=("TKDefaultFont", 12))
        client_label.grid(row=0)


    @staticmethod
    def _on_quit():
        """Command to exit program"""
        sys.exit()


class MainPage(AppPage):
    """Main page to choose whether to create, update, or run report."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        m_menu = self._add_frame("Main Menu")
        self.create_button = ttk.Button(
            m_menu,
            text='Create Project',
            command=self._on_create
        )
        self.create_button.grid(row=0, column=0, sticky='ew')

        self.update_button = ttk.Button(
            m_menu,
            text='Update Project',
            command=self._on_update
        )
        self.update_button.grid(row=1, column=0, sticky='ew')

        self.report_button = ttk.Button(
            m_menu,
            text='Run Report',
            command=self._on_report
        )
        self.report_button.grid(row=2, column=0, sticky='ew')

        self.quit_button = ttk.Button(
            m_menu,
            text='Quit',
            command=self._on_quit
        )
        self.quit_button.grid(row=3, column=0, sticky='ew')

    @staticmethod
    def _on_create():
        """Command to launch new project page"""
        app.create_page1()

    def _on_update(self):
        """Command to update existing project"""
        app.update_page1()

    def _on_report(self):
        """Command to open report page"""
        pass


# noinspection PyTypeChecker
class CreatePage1(AppPage):
    """Project create page #1"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)


        c_info = self._add_frame("Client Information")

        LabelInput(
            c_info, "Client Name", var=self._vars['Client Name']
        ).grid(row=0, column=0)
        self.cal = Calendar(c_info, selectmode='day', year=2023)
        self.cal.grid(row=2, column=0)
        self.date_label = tk.Label(
            c_info,
            text='Choose Go Live / Dead Date',
        )
        self.date_label.grid(row=1, column=0)

        m_info = self._add_frame('Market')

        LabelInput(
            m_info, "", input_class=ttk.Radiobutton,
            var=self._vars['Market'],
            input_args={"values": ["NO", "SHV", "ATX",
                                   "IND", "VAL", "KC",
                                   "DFW", "HOU"]},
        ).grid(row=0, column=0)

        t_info = self._add_frame('Project Type')

        LabelInput(
            t_info, "", input_class=ttk.Radiobutton,
            var=self._vars['Type'],
            input_args={"values": ["Onboarding", "Offboarding", "Conversion"]}
        ).grid(row=0, column=0)

        buttons = tk.Frame(self)
        buttons.grid(sticky=tk.W + tk.E + tk.S, row=3)

        self.quit_button = ttk.Button(
            buttons, text="Quit", command=self._on_quit
        )
        self.quit_button.pack(side=tk.RIGHT)
        self.next_button = ttk.Button(
            buttons, text="Next", command=self._on_create2
        )
        self.next_button.pack(side=tk.RIGHT)

    def _on_create2(self):
        """Move onto creation page two
        (which will be one of three pages depending on Type selected"""
        self._vars['Go Live / Dead Date'].set(self.cal.get_date())
        app.create_page2()


# noinspection PyTypeChecker
class CreatePage2On(AppPage):
    """Creation Page 2 if 'Onboardings' is type"""

    def __init__(self, client, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.create_title(client, 'Onboarding')

        self._tpg_tools('Onboarding', 'AV', 'MDR', 'Barracuda', 'Ninjio')

        buttons = tk.Frame(self)
        buttons.grid(sticky=tk.W + tk.E + tk.S, row=5)

        self.quit_button = ttk.Button(
            buttons, text="Quit", command=self._on_quit
        )
        self.quit_button.pack(side=tk.RIGHT)
        self.next_button = ttk.Button(
            buttons, text="Save", command=self._on_create2on_save
        )
        self.next_button.pack(side=tk.RIGHT)

    def _on_create2on_save(self):
        app.creation_complete()


class CreatePage2Off(AppPage):
    """Creation Page 2 if 'Offboardings' is type"""

    def __init__(self, client, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.create_title(client, 'Offboarding')

        self._tpg_tools('Offboarding', 'AV', 'MDR', 'Barracuda', 'Ninjio')

        buttons = tk.Frame(self)
        buttons.grid(sticky=tk.W + tk.E + tk.S, row=5)

        self.quit_button = ttk.Button(
            buttons, text="Quit", command=self._on_quit
        )
        self.quit_button.pack(side=tk.RIGHT)
        self.next_button = ttk.Button(
            buttons, text="Save", command=self._on_createoff_save
        )
        self.next_button.pack(side=tk.RIGHT)

    def _on_createoff_save(self):
        app.creation_complete()


class CreatePage2Conv(AppPage):
    """Creation Page 2 if 'Conversions' is type"""

    def __init__(self, client, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.create_title(client, 'Conversion')
        self.old_label = ttk.Label(
            self,
            text="Old Plan Tools",
            font=("TKDefaultFont", 11))
        self.old_label.grid(row=1)
        self._tpg_tools('Conversion', 'AV', 'MDR', 'Barracuda', 'Ninjio')
        self.old_label = ttk.Label(
            self,
            text="New Plan Tools",
            font=("TKDefaultFont", 11))
        self.old_label.grid(row=6)
        self._tpg_tools('Onboarding',
                        'AV-conv',
                        'MDR-conv',
                        'Barracuda-conv',
                        'Ninjio-conv',
                        )

        buttons = tk.Frame(self)
        buttons.grid(sticky=tk.W + tk.E + tk.S, row=99)

        self.quit_button = ttk.Button(
            buttons, text="Quit", command=self._on_quit
        )
        self.quit_button.pack(side=tk.RIGHT)
        self.next_button = ttk.Button(
            buttons, text="Save", command=self._on_createconv_save
        )
        self.next_button.pack(side=tk.RIGHT)

    def _on_createconv_save(self):
        app.creation_complete()


class UpdatePage1(AppPage):
    """Page 1 of update section to choose client"""

    def __init__(self, client_lists, *args, **kwargs):
        super().__init__(*args, **kwargs)

        tk.Label(self, text='Choose Client Project to update'
                 ).grid(row=0, column=0, sticky=(tk.W + tk.E))
        self.client_choice = []
        for i in client_lists:
            if i[0] == 'Client Name':
                continue
            else:
                client_proj = f'{i[2]} - {i[0]} - {i[3]}'
                self.client_choice.append(client_proj)
        self.client_choice = sorted(self.client_choice)
        self.max_len = 0
        # expand width by longest client name in list
        for i in self.client_choice:
            if len(i) > self.max_len:
                self.max_len = len(i)
        self.client_box = tk.Listbox(self, height=20, width=self.max_len)
        self.client_box.grid(row=1, column=0)
        self.client_box.delete(0, 'end')
        for i in self.client_choice:
            self.client_box.insert('end', i)

        buttons = tk.Frame(self)
        buttons.grid(sticky=tk.W + tk.E + tk.S, row=99)

        self.quit_button = ttk.Button(
            buttons, text="Quit", command=self._on_quit
        )
        self.quit_button.pack(side=tk.RIGHT)
        self.next_button = ttk.Button(
            buttons, text="Next", command=self._on_update1_next
        )
        self.next_button.pack(side=tk.RIGHT)

    def _on_update1_next(self):
        choice = ''
        for i in self.client_box.curselection():
             choice = self.client_box.get(i)
        self._vars['update1_choice'] = tk.StringVar(value=choice)
        app.update_page2()


class UpdatePage2(AppPage):
    """Page 2 of the update page. Different layout depending on project type"""

    def __init__(self, client_proj_list, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.create_title(client_proj_list[0], client_proj_list[3])
        self.task_frame = tk.Frame()
        # if client_proj_list[3] == 'Onboarding':
            # LabelInput(
            #     ninjio_info, "Ninjio", input_class=ttk.Checkbutton,
            #     var=self._vars[ninjio],
            # ).grid(row=0, column=0)
            #




class Application(tk.Tk):
    """Application root window"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.title("TPG Onboarding Project Companion")
        self.columnconfigure(0, weight=1)
        self.minsize(400, 300)
        self.main_label = ''
        self.m_page = ''
        self.c_page1 = CreatePage1(self)
        self.c_page2_on = CreatePage2On(self)
        self.c_page2_off = CreatePage2Off(self)
        self.c_page2_conv = CreatePage2Conv(self)
        self.client_data = {}
        self.plan_data = {}
        self.client_projects = []
        self.u_page1 = UpdatePage1(self.client_projects)
        self.update1_choice_client = ''
        self.u_page2 = UpdatePage2('Failed to pull client data', 'Try again')


        self.status = tk.StringVar()
        ttk.Label(self, textvariable=self.status
                  ).grid(sticky=(tk.W + tk.E), row=2, padx=10)
        self.main_page()
        self.db_file = 'ProjectCompanion-DBv1.xlsx'
        self.update1_choice = ''

    def client_check(self, wb_sheet, client):
        """
        Pass the needed sheet name and device name.
        Iterate through the sheet and look for the existence of the client name.
        If the client is present, return the row it resides in,
        if the client is not present return empty value for device.
        Needed modules: openpyxl, Workbook
        :param wb_sheet: The sheet to check
        :param client: The client name to search for
        :return: The row number the device resides in.
        """
        max_row = wb_sheet.max_row
        client_row = ''
        for i in range(1, max_row + 1):
            cell_data = wb_sheet.cell(row=i, column=1).value
            if client in cell_data:
                client_row = i
                break
            else:
                client_row = ''
        return client_row

    def main_page(self):
        self.m_page = MainPage(self)
        self.main_label = ttk.Label(
            self,
            text="TPG Onboarding Project Companion",
            font=("TKDefaultFont", 14))
        self.main_label.grid(row=0)
        self.m_page.grid(row=1, padx=10, sticky=(tk.W + tk.E))

    def create_page1(self):
        self.main_label.grid_forget()
        self.main_label.destroy()
        self.m_page.grid_forget()
        self.m_page.destroy()
        self.c_page1 = CreatePage1(self)
        self.c_page1.grid(row=1, padx=10, sticky=(tk.W + tk.E))
        self.status.set(str(''))

    def create_page2(self):
        try:
            self.client_data = self.c_page1.get()
        except ValueError as e:
            self.status.set(str(e))
            return
        client_name = self.client_data['Client Name']

        if self.client_data['Market'] == '':
            self.status.set('Required Market Location is Missing!')

        self.c_page1.grid_forget()
        self.c_page1.destroy()

        if self.client_data['Type'] == 'Onboarding':
            self.c_page2_on = CreatePage2On(client_name)
            self.c_page2_on.grid(row=1, padx=10, sticky=(tk.W + tk.E))
        elif self.client_data['Type'] == 'Offboarding':
            self.c_page2_off = CreatePage2Off(client_name)
            self.c_page2_off.grid(row=1, padx=10, sticky=(tk.W + tk. E))
        elif self.client_data['Type'] == 'Conversion':
            self.c_page2_conv = CreatePage2Conv(client_name)
            self.c_page2_conv.grid(row=1, padx=10, sticky=(tk.W + tk.E))
        else:
            self.status.set('Required Project Type is Missing!')

    def project_creation(self):
        """Method used by all create project pages to write new project to db"""
        wb = load_workbook(self.db_file)
        sheet = wb['Sheet1']
        client_row = self.client_check(sheet, self.client_data['Client Name'])
        if client_row == '':
            client_row = sheet.max_row + 1

            new_row = [(self.client_data['Client Name'],
                        self.client_data['Go Live / Dead Date'],
                        self.client_data['Market'],
                        self.client_data['Type'],
                        self.plan_data['AV'],
                        self.plan_data['MDR'],
                        self.plan_data['Ninjio'],
                        self.plan_data['Barracuda'],
                        self.plan_data['AV-conv'],
                        self.plan_data['MDR-conv'],
                        self.plan_data['Ninjio-conv'],
                        self.plan_data['Barracuda-conv'],
                        )]

            for row in new_row:
                sheet.append(row)
        else:
            self.status.set(str(f"A project already exists for "
                                f"{self.client_data['Client Name']}! "
                                f"Data not saved!"))
        wb.save(self.db_file)
        self.status.set(str(f"A project has been created for "
                            f"{self.client_data['Client Name']}"))

    def creation_complete(self):
        try:
            for i in (self.c_page2_on, self.c_page2_off, self.c_page2_conv):
                current_frame = ''
                if i.winfo_exists():
                    current_frame = i
                    break
        except AttributeError:
            return

        try:
            self.plan_data = current_frame.get()
        except ValueError as e:
            self.status.set(str(e))
            return
        current_frame.grid_forget()
        current_frame.destroy()
        self.main_page()
        self.project_creation()


    def pull_dbv1(self, wb, sheet):
        wb = load_workbook(wb)
        sheet = wb[sheet]
        proj_list = []
        for row in sheet.rows:
            temp_list = []
            for item in row:
                temp_list.append(item.value)
            proj_list.append(temp_list)
        wb.close()
        return proj_list


    def update_page1(self):
        self.main_label.grid_forget()
        self.main_label.destroy()
        self.m_page.grid_forget()
        self.m_page.destroy()
        self.client_projects = self.pull_dbv1(self.db_file, 'Sheet1')
        self.u_page1 = UpdatePage1(self.client_projects)
        self.u_page1.grid(row=1, padx=10, sticky=(tk.W + tk. E))

    def update_page2(self):
        try:
            self.update1_choice = self.u_page1.get()
        except ValueError as e:
            self.status.set(str(e))
            return
        self.update1_choice_list = \
            self.update1_choice['update1_choice'].split(' - ')
        self.update1_choice_client = self.update1_choice_list[1]
        proj_list = self.pull_dbv1(self.db_file, 'Sheet1')
        for i in proj_list:
            if i[0] == self.update1_choice_client:
                client_proj_data = i
        print(client_proj_data)
        self.u_page1.grid_forget()
        self.u_page1.destroy()
        self.u_page2 = UpdatePage2(client_proj_data)



if __name__ == "__main__":
    app = Application()
    app.mainloop()